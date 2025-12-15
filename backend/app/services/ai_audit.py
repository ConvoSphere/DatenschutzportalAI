import os
import logging
import asyncio
from typing import List, Dict, Any, Optional
from datetime import datetime
import json
from pathlib import Path

# Text extraction libraries
import pypdf
import docx
import openpyxl
from odf.opendocument import load as load_odf
from odf.text import P

# AI libraries
from pydantic_ai import Agent, RunContext
from pydantic import BaseModel, Field
from app.config import settings
from app.config.audit_criteria import DEFAULT_AUDIT_CRITERIA, CheckItem

# Nextcloud service
from app.services.nextcloud import NextcloudService

logger = logging.getLogger(__name__)

# --- Data Models for AI Output ---

class CheckResult(BaseModel):
    check_id: str = Field(description="ID of the check item from criteria")
    status: str = Field(description="Status of the check: 'PASS', 'FAIL', 'WARNING', 'UNKNOWN'")
    findings: str = Field(description="Detailed findings and observations")
    recommendation: Optional[str] = Field(description="Recommendation for improvement if applicable")

class AuditResult(BaseModel):
    summary: str = Field(description="High-level executive summary of the audit")
    results: List[CheckResult] = Field(description="Detailed results for each check item")
    overall_status: str = Field(description="Overall status: 'PASS', 'NEEDS_IMPROVEMENT', 'FAIL'")

# --- Service Class ---

class AIAuditService:
    def __init__(self):
        self.nextcloud = NextcloudService()
        self.criteria = DEFAULT_AUDIT_CRITERIA
        
        # Set OpenAI environment variables for Pydantic AI (which uses OpenAI SDK)
        # This ensures that if the user configured a custom OpenAI-compatible endpoint, it is used.
        if settings.ai_api_key:
            os.environ["OPENAI_API_KEY"] = settings.ai_api_key
        if settings.ai_api_base_url:
            os.environ["OPENAI_BASE_URL"] = settings.ai_api_base_url
        
        # Configure proxy if provided (useful for corporate networks or tunneling)
        if settings.ai_proxy:
             # Standard proxy env vars for requests/aiohttp/httpx
            os.environ["HTTP_PROXY"] = settings.ai_proxy
            os.environ["HTTPS_PROXY"] = settings.ai_proxy
            # OpenAI specific (though usually standard env vars are enough)
            # os.environ["OPENAI_PROXY"] = settings.ai_proxy 
        
        self.agent = Agent(
            model=settings.ai_model_name,
            system_prompt=self.criteria.system_prompt,
            result_type=AuditResult,
        )

    async def perform_audit(self, project_id: str, file_paths: List[str]) -> AuditResult:
        """
        Main entry point for the audit process.
        :param project_id: The ID of the project in Nextcloud
        :param file_paths: List of temporary local paths to the files (or downloaded files)
        """
        try:
            logger.info(f"Starting AI audit for project {project_id} with {len(file_paths)} files")
            
            # 1. Extract text from all files
            combined_text = ""
            for file_path in file_paths:
                text = self._extract_text(file_path)
                filename = os.path.basename(file_path)
                if text:
                    combined_text += f"\n\n--- FILE: {filename} ---\n\n{text}"
                else:
                    logger.warning(f"Could not extract text from {filename}")
                    combined_text += f"\n\n--- FILE: {filename} (Text extraction failed or empty) ---\n\n"

            if not combined_text.strip():
                logger.error("No text could be extracted from any file.")
                return AuditResult(
                    summary="Audit failed because no text could be extracted from the uploaded documents.",
                    results=[],
                    overall_status="FAIL"
                )

            # 2. Run AI Analysis
            # We inject the check items into the prompt context if needed, or rely on system prompt.
            # Ideally, pass them as context or part of the user prompt.
            
            user_prompt = f"""
            Bitte prüfe die folgenden Dokumenteninhalte gegen die Checkliste.
            
            Checkliste:
            {json.dumps([item.model_dump() for item in self.criteria.check_items], indent=2)}
            
            Dokumenteninhalte:
            {combined_text[:100000]} 
            """
            # Truncate text if too long (simple protection, though model context might be larger)
            # 100k chars is approx 25k tokens.
            
            logger.info("Running AI analysis...")
            result = await self.agent.run(user_prompt)
            audit_result = result.data
            
            logger.info(f"AI analysis completed. Status: {audit_result.overall_status}")
            
            return audit_result

        except Exception as e:
            logger.error(f"Audit failed: {e}", exc_info=True)
            return AuditResult(
                summary=f"An error occurred during the automated audit: {str(e)}",
                results=[],
                overall_status="FAIL"
            )

    def _extract_text(self, file_path: str) -> str:
        """Extract text based on file extension."""
        ext = os.path.splitext(file_path)[1].lower()
        try:
            if ext == '.pdf':
                return self._extract_from_pdf(file_path)
            elif ext in ['.docx', '.doc']:
                return self._extract_from_docx(file_path)
            elif ext in ['.xlsx', '.xls']:
                return self._extract_from_excel(file_path)
            elif ext == '.odt':
                return self._extract_from_odt(file_path)
            elif ext in ['.txt', '.md']:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    return f.read()
            else:
                logger.warning(f"Unsupported file type for text extraction: {ext}")
                return ""
        except Exception as e:
            logger.error(f"Error extracting text from {file_path}: {e}")
            return ""

    def _extract_from_pdf(self, path: str) -> str:
        text = ""
        with open(path, 'rb') as f:
            reader = pypdf.PdfReader(f)
            for page in reader.pages:
                text += page.extract_text() + "\n"
        return text

    def _extract_from_docx(self, path: str) -> str:
        doc = docx.Document(path)
        return "\n".join([p.text for p in doc.paragraphs])

    def _extract_from_excel(self, path: str) -> str:
        wb = openpyxl.load_workbook(path, data_only=True)
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text += f"\nSheet: {sheet}\n"
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) for cell in row if cell is not None])
                if row_text:
                    text += row_text + "\n"
        return text

    def _extract_from_odt(self, path: str) -> str:
        doc = load_odf(path)
        text = []
        for p in doc.getElementsByType(P):
            text.append(str(p)) # This might need refinement for ODT text extraction
            # odfpy is a bit tricky, alternative is teletype
        
        # Better simple extraction for odfpy
        from odf import teletype
        return "\n".join([teletype.extractText(p) for p in doc.getElementsByType(P)])

    async def generate_report(self, audit_result: AuditResult, output_path: str):
        """Generates a Markdown report and saves it."""
        
        md = f"# Automatische Datenschutz-Prüfung\n\n"
        md += f"**Datum:** {datetime.now().strftime('%d.%m.%Y %H:%M')}\n"
        md += f"**Gesamtergebnis:** {audit_result.overall_status}\n\n"
        
        md += "## Zusammenfassung\n\n"
        md += f"{audit_result.summary}\n\n"
        
        md += "## Detaillierte Prüfung\n\n"
        
        # Group by category if possible, or just list
        for result in audit_result.results:
            icon = "✅" if result.status == 'PASS' else "⚠️" if result.status == 'WARNING' else "❌" if result.status == 'FAIL' else "❓"
            md += f"### {icon} {result.check_id}\n"
            md += f"**Status:** {result.status}\n\n"
            md += f"**Befund:**\n{result.findings}\n\n"
            if result.recommendation:
                md += f"**Empfehlung:**\n{result.recommendation}\n\n"
            md += "---\n\n"
            
        md += "\n*Hinweis: Dieser Bericht wurde automatisch durch KI erstellt und dient als Unterstützung. Er ersetzt keine rechtliche Prüfung durch einen Datenschutzbeauftragten.*\n"
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(md)
            
        return md

